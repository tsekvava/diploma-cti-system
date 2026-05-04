"""
MITRE ATT&CK Catalog — парсер XLSX (v18 перевод) и STIX JSON в SQLite.

Создаёт локальную базу mitre.db с таблицами:
  - tactics       (14 тактик с RU/EN названиями и описаниями)
  - techniques     (692 техники/подтехники с RU/EN названиями и описаниями)
  - groups         (APT-группы из STIX, с алиасами)
  - software       (малварь + инструменты из STIX, с алиасами)
  - group_techniques   (связи группа → техника)
  - software_techniques (связи software → техника)
  - data_sources   (источники данных для детектирования)
  - mitigations    (меры предотвращения)

Использование:
    python -m knowledge_base.mitre_catalog          # Парсит XLSX
    python -m knowledge_base.mitre_catalog --stix   # Парсит XLSX + STIX JSON
"""

import sqlite3
import json
import os
import re
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Установи openpyxl: pip install openpyxl")
    raise

PROJECT_ROOT = Path(__file__).parent.parent
XLSX_PATH = PROJECT_ROOT / "data" / "mitre" / "mitre_v18_translation.xlsm"
STIX_PATH = PROJECT_ROOT / "data" / "mitre" / "enterprise-attack.json"
DB_PATH = PROJECT_ROOT / "knowledge_base" / "mitre.db"

# 14 тактик MITRE ATT&CK Enterprise в порядке Kill Chain
TACTIC_ORDER = {
    "TA0043": ("Reconnaissance", "Разведка", 1),
    "TA0042": ("Resource Development", "Подготовка ресурсов", 2),
    "TA0001": ("Initial Access", "Первоначальный доступ", 3),
    "TA0002": ("Execution", "Выполнение", 4),
    "TA0003": ("Persistence", "Закрепление", 5),
    "TA0004": ("Privilege Escalation", "Повышение привилегий", 6),
    "TA0005": ("Defense Evasion", "Предотвращение обнаружения", 7),
    "TA0006": ("Credential Access", "Получение учетных данных", 8),
    "TA0007": ("Discovery", "Обнаружение", 9),
    "TA0008": ("Lateral Movement", "Перемещение внутри периметра", 10),
    "TA0009": ("Collection", "Сбор данных", 11),
    "TA0011": ("Command and Control", "Управление и контроль", 12),
    "TA0010": ("Exfiltration", "Эксфильтрация", 13),
    "TA0040": ("Impact", "Деструктивное воздействие", 14),
}


def create_database(db_path: Path) -> sqlite3.Connection:
    """Создаёт SQLite базу со всеми таблицами."""
    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")

    conn.executescript("""
        CREATE TABLE IF NOT EXISTS tactics (
            id TEXT PRIMARY KEY,
            name_en TEXT NOT NULL,
            name_ru TEXT NOT NULL,
            description_ru TEXT DEFAULT '',
            kill_chain_order INTEGER NOT NULL
        );

        CREATE TABLE IF NOT EXISTS techniques (
            id TEXT PRIMARY KEY,
            name_en TEXT NOT NULL DEFAULT '',
            name_ru TEXT NOT NULL,
            description_ru TEXT DEFAULT '',
            tactic_id TEXT,
            parent_technique_id TEXT,
            is_subtechnique BOOLEAN DEFAULT 0,
            FOREIGN KEY (tactic_id) REFERENCES tactics(id)
        );

        CREATE TABLE IF NOT EXISTS groups (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            description TEXT DEFAULT '',
            aliases TEXT DEFAULT '[]',
            country TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS software (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            type TEXT DEFAULT 'malware',
            description TEXT DEFAULT '',
            aliases TEXT DEFAULT '[]'
        );

        CREATE TABLE IF NOT EXISTS group_techniques (
            group_id TEXT,
            technique_id TEXT,
            PRIMARY KEY (group_id, technique_id),
            FOREIGN KEY (group_id) REFERENCES groups(id),
            FOREIGN KEY (technique_id) REFERENCES techniques(id)
        );

        CREATE TABLE IF NOT EXISTS software_techniques (
            software_id TEXT,
            technique_id TEXT,
            PRIMARY KEY (software_id, technique_id),
            FOREIGN KEY (software_id) REFERENCES software(id),
            FOREIGN KEY (technique_id) REFERENCES techniques(id)
        );

        CREATE TABLE IF NOT EXISTS data_sources (
            id TEXT PRIMARY KEY,
            name_en TEXT DEFAULT '',
            name_ru TEXT DEFAULT '',
            description_ru TEXT DEFAULT '',
            components TEXT DEFAULT '[]'
        );

        CREATE TABLE IF NOT EXISTS mitigations (
            id TEXT PRIMARY KEY,
            name_en TEXT DEFAULT '',
            name_ru TEXT DEFAULT '',
            description_ru TEXT DEFAULT '',
            techniques_covered TEXT DEFAULT '[]'
        );

        CREATE INDEX IF NOT EXISTS idx_techniques_tactic ON techniques(tactic_id);
        CREATE INDEX IF NOT EXISTS idx_techniques_parent ON techniques(parent_technique_id);
        CREATE INDEX IF NOT EXISTS idx_group_techniques_group ON group_techniques(group_id);
        CREATE INDEX IF NOT EXISTS idx_software_techniques_sw ON software_techniques(software_id);
    """)
    return conn


def parse_xlsx_tactics_and_techniques(conn: sqlite3.Connection, xlsx_path: Path):
    """Парсит лист 'Enterprise ATT&CK (v16)' — тактики, техники, подтехники."""
    print(f"[1/4] Открываю XLSX: {xlsx_path.name}...")
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)

    # --- Тактики ---
    print("   Загружаю тактики...")
    ws = wb["Enterprise ATT&CK (v16)"]
    rows = list(ws.iter_rows(values_only=True))

    # Строка 0: ID тактик (TA0043, TA0042, ...) через каждые 4 колонки
    tactic_ids_row = rows[0]
    # Строка 2: Описания тактик
    tactic_desc_row = rows[2] if len(rows) > 2 else [None] * len(tactic_ids_row)

    tactic_columns = {}  # {column_index: tactic_id}

    for col_idx in range(0, len(tactic_ids_row), 4):
        tactic_id = tactic_ids_row[col_idx]
        if tactic_id and str(tactic_id).startswith("TA"):
            tactic_id = str(tactic_id).strip()
            desc_ru = str(tactic_desc_row[col_idx] or "") if col_idx < len(tactic_desc_row) else ""

            if tactic_id in TACTIC_ORDER:
                name_en, name_ru, order = TACTIC_ORDER[tactic_id]
                conn.execute(
                    "INSERT OR REPLACE INTO tactics VALUES (?, ?, ?, ?, ?)",
                    (tactic_id, name_en, name_ru, desc_ru, order)
                )
                tactic_columns[col_idx] = tactic_id

    tactic_count = conn.execute("SELECT COUNT(*) FROM tactics").fetchone()[0]
    print(f"   -> {tactic_count} тактик загружено")

    # --- Техники и подтехники ---
    print("   Загружаю техники и подтехники...")
    technique_count = 0
    sub_count = 0

    for row in rows[3:]:  # Данные начинаются с 4-й строки
        for col_idx, tactic_id in tactic_columns.items():
            # Колонки: [tech_id, tech_name_ru, sub_id, sub_name_ru]
            tech_id = row[col_idx] if col_idx < len(row) else None
            tech_name = row[col_idx + 1] if col_idx + 1 < len(row) else None
            sub_id = row[col_idx + 2] if col_idx + 2 < len(row) else None
            sub_name = row[col_idx + 3] if col_idx + 3 < len(row) else None

            if tech_id and str(tech_id).startswith("T"):
                tech_id = str(tech_id).strip()
                tech_name = str(tech_name or "").strip()
                conn.execute(
                    """INSERT OR IGNORE INTO techniques
                       (id, name_ru, tactic_id, is_subtechnique, parent_technique_id)
                       VALUES (?, ?, ?, 0, NULL)""",
                    (tech_id, tech_name, tactic_id)
                )
                technique_count += 1

            if sub_id and str(sub_id).startswith("T"):
                sub_id = str(sub_id).strip()
                sub_name = str(sub_name or "").strip()
                parent_id = sub_id.split(".")[0]
                conn.execute(
                    """INSERT OR IGNORE INTO techniques
                       (id, name_ru, tactic_id, is_subtechnique, parent_technique_id)
                       VALUES (?, ?, ?, 1, ?)""",
                    (sub_id, sub_name, tactic_id, parent_id)
                )
                sub_count += 1

    total = conn.execute("SELECT COUNT(*) FROM techniques").fetchone()[0]
    print(f"   -> {total} записей (техники + подтехники)")

    # --- Описания из листа v18 ---
    print("[2/4] Загружаю описания из 'Перевод описаний v18'...")
    ws_desc = wb["Перевод описаний v18"]
    desc_rows = list(ws_desc.iter_rows(values_only=True))

    desc_updated = 0
    for row in desc_rows[4:]:  # Данные с 5-й строки (после заголовков)
        # Каждая тактика занимает 6 колонок: tech_id, tech_name, tech_desc, sub_id, sub_name, sub_desc
        for tactic_start in range(0, min(len(row), 84), 6):
            tech_id = row[tactic_start] if tactic_start < len(row) else None
            tech_desc = row[tactic_start + 2] if tactic_start + 2 < len(row) else None
            sub_id = row[tactic_start + 3] if tactic_start + 3 < len(row) else None
            sub_desc = row[tactic_start + 5] if tactic_start + 5 < len(row) else None

            if tech_id and str(tech_id).startswith("T") and tech_desc:
                conn.execute(
                    "UPDATE techniques SET description_ru = ? WHERE id = ?",
                    (str(tech_desc).strip(), str(tech_id).strip())
                )
                desc_updated += 1

            if sub_id and str(sub_id).startswith("T") and sub_desc:
                conn.execute(
                    "UPDATE techniques SET description_ru = ? WHERE id = ?",
                    (str(sub_desc).strip(), str(sub_id).strip())
                )
                desc_updated += 1

    print(f"   -> {desc_updated} описаний обновлено")

    # --- Источники данных ---
    print("[3/4] Загружаю источники данных...")
    ws_ds = wb["Источники данных"]
    ds_rows = list(ws_ds.iter_rows(values_only=True))

    ds_count = 0
    current_ds_id = None
    components = []

    for row in ds_rows[2:]:  # Пропускаем заголовки
        ds_id = row[0] if row[0] else None
        name_en = row[1] if len(row) > 1 else None
        name_ru = row[2] if len(row) > 2 else None
        desc_ru = row[3] if len(row) > 3 else None
        comp_name = row[4] if len(row) > 4 else None

        if ds_id and str(ds_id).startswith("DS"):
            # Сохранить предыдущий DS
            if current_ds_id:
                conn.execute(
                    "INSERT OR REPLACE INTO data_sources VALUES (?, ?, ?, ?, ?)",
                    (current_ds_id, current_name_en or "", current_name_ru or "",
                     current_desc_ru or "", json.dumps(components, ensure_ascii=False))
                )
                ds_count += 1

            current_ds_id = str(ds_id).strip()
            current_name_en = str(name_en or "").strip()
            current_name_ru = str(name_ru or "").strip()
            current_desc_ru = str(desc_ru or "").strip()
            components = []

        if comp_name and current_ds_id:
            components.append(str(comp_name).strip())

    # Последний DS
    if current_ds_id:
        conn.execute(
            "INSERT OR REPLACE INTO data_sources VALUES (?, ?, ?, ?, ?)",
            (current_ds_id, current_name_en or "", current_name_ru or "",
             current_desc_ru or "", json.dumps(components, ensure_ascii=False))
        )
        ds_count += 1

    print(f"   -> {ds_count} источников данных загружено")

    # --- Меры предотвращения ---
    print("[4/4] Загружаю меры предотвращения...")
    ws_mit = wb["Предотвращение"]
    mit_rows = list(ws_mit.iter_rows(values_only=True))

    mit_count = 0
    current_mit_id = None
    covered_techniques = []

    for row in mit_rows[2:]:
        mit_id = row[0] if row[0] else None
        name_en = row[1] if len(row) > 1 else None
        name_ru = row[2] if len(row) > 2 else None
        desc_ru = row[3] if len(row) > 3 else None
        tech_covered = row[4] if len(row) > 4 else None

        if mit_id and str(mit_id).startswith("M"):
            if current_mit_id:
                conn.execute(
                    "INSERT OR REPLACE INTO mitigations VALUES (?, ?, ?, ?, ?)",
                    (current_mit_id, current_m_name_en or "", current_m_name_ru or "",
                     current_m_desc_ru or "", json.dumps(covered_techniques, ensure_ascii=False))
                )
                mit_count += 1

            current_mit_id = str(mit_id).strip()
            current_m_name_en = str(name_en or "").strip()
            current_m_name_ru = str(name_ru or "").strip()
            current_m_desc_ru = str(desc_ru or "").strip()
            covered_techniques = []

        if tech_covered and current_mit_id:
            # Извлечь T-коды из строки вида "T1110.001, T1110.003"
            t_codes = re.findall(r"T\d{4}(?:\.\d{3})?", str(tech_covered))
            covered_techniques.extend(t_codes)

    if current_mit_id:
        conn.execute(
            "INSERT OR REPLACE INTO mitigations VALUES (?, ?, ?, ?, ?)",
            (current_mit_id, current_m_name_en or "", current_m_name_ru or "",
             current_m_desc_ru or "", json.dumps(covered_techniques, ensure_ascii=False))
        )
        mit_count += 1

    print(f"   -> {mit_count} мер предотвращения загружено")

    wb.close()
    conn.commit()


def parse_stix_json(conn: sqlite3.Connection, stix_path: Path):
    """Парсит STIX enterprise-attack.json — APT-группы, software, связи."""
    if not stix_path.exists():
        print(f"\n[STIX] Файл не найден: {stix_path}")
        print("   Скачай его: https://github.com/mitre/cti/raw/master/enterprise-attack/enterprise-attack.json")
        print(f"   Положи в: {stix_path}")
        return

    print(f"\n[STIX] Загружаю {stix_path.name}...")
    with open(stix_path, "r", encoding="utf-8") as f:
        bundle = json.load(f)

    objects = bundle.get("objects", [])
    print(f"   -> {len(objects)} STIX-объектов в файле")

    # Маппинг STIX ID → MITRE ID
    stix_to_mitre = {}

    # --- Группы ---
    group_count = 0
    for obj in objects:
        if obj.get("type") != "intrusion-set":
            continue
        if obj.get("revoked", False) or obj.get("x_mitre_deprecated", False):
            continue

        refs = obj.get("external_references", [])
        mitre_id = None
        for ref in refs:
            if ref.get("source_name") == "mitre-attack":
                mitre_id = ref.get("external_id")
                break

        if not mitre_id:
            continue

        stix_to_mitre[obj["id"]] = mitre_id
        aliases = obj.get("aliases", [])
        name = obj.get("name", "")
        # Убираем имя из алиасов (STIX дублирует)
        aliases = [a for a in aliases if a != name]

        conn.execute(
            "INSERT OR REPLACE INTO groups VALUES (?, ?, ?, ?, ?)",
            (mitre_id, name, obj.get("description", "")[:2000],
             json.dumps(aliases, ensure_ascii=False), "")
        )
        group_count += 1

    print(f"   -> {group_count} APT-групп загружено")

    # --- Software (malware + tool) ---
    sw_count = 0
    for obj in objects:
        if obj.get("type") not in ("malware", "tool"):
            continue
        if obj.get("revoked", False) or obj.get("x_mitre_deprecated", False):
            continue

        refs = obj.get("external_references", [])
        mitre_id = None
        for ref in refs:
            if ref.get("source_name") == "mitre-attack":
                mitre_id = ref.get("external_id")
                break

        if not mitre_id:
            continue

        stix_to_mitre[obj["id"]] = mitre_id
        aliases = obj.get("x_mitre_aliases", [])
        name = obj.get("name", "")
        aliases = [a for a in aliases if a != name]

        conn.execute(
            "INSERT OR REPLACE INTO software VALUES (?, ?, ?, ?, ?)",
            (mitre_id, name, obj["type"],
             obj.get("description", "")[:2000],
             json.dumps(aliases, ensure_ascii=False))
        )
        sw_count += 1

    print(f"   -> {sw_count} software записей (malware + tools)")

    # --- Technique STIX ID mapping ---
    for obj in objects:
        if obj.get("type") != "attack-pattern":
            continue
        refs = obj.get("external_references", [])
        for ref in refs:
            if ref.get("source_name") == "mitre-attack":
                stix_to_mitre[obj["id"]] = ref.get("external_id")
                # Также обновляем name_en в techniques
                mitre_id = ref.get("external_id")
                name_en = obj.get("name", "")
                conn.execute(
                    "UPDATE techniques SET name_en = ? WHERE id = ?",
                    (name_en, mitre_id)
                )
                break

    # --- Связи (relationships) ---
    rel_group_tech = 0
    rel_sw_tech = 0

    for obj in objects:
        if obj.get("type") != "relationship":
            continue
        if obj.get("revoked", False):
            continue
        if obj.get("relationship_type") != "uses":
            continue

        source_id = stix_to_mitre.get(obj.get("source_ref"))
        target_id = stix_to_mitre.get(obj.get("target_ref"))

        if not source_id or not target_id:
            continue

        # Группа → техника
        if source_id.startswith("G") and target_id.startswith("T"):
            # Проверяем что техника существует в базе
            exists = conn.execute(
                "SELECT 1 FROM techniques WHERE id = ?", (target_id,)
            ).fetchone()
            if exists:
                conn.execute(
                    "INSERT OR IGNORE INTO group_techniques VALUES (?, ?)",
                    (source_id, target_id)
                )
                rel_group_tech += 1

        # Software → техника
        if source_id.startswith("S") and target_id.startswith("T"):
            exists = conn.execute(
                "SELECT 1 FROM techniques WHERE id = ?", (target_id,)
            ).fetchone()
            if exists:
                conn.execute(
                    "INSERT OR IGNORE INTO software_techniques VALUES (?, ?)",
                    (source_id, target_id)
                )
                rel_sw_tech += 1

    print(f"   -> {rel_group_tech} связей группа→техника")
    print(f"   -> {rel_sw_tech} связей software→техника")

    conn.commit()


def print_stats(conn: sqlite3.Connection):
    """Выводит статистику по базе."""
    print("\n" + "=" * 50)
    print("СТАТИСТИКА БАЗЫ MITRE ATT&CK")
    print("=" * 50)

    tables = [
        ("tactics", "Тактик"),
        ("techniques", "Техник/подтехник"),
        ("groups", "APT-групп"),
        ("software", "Software (malware + tools)"),
        ("group_techniques", "Связей группа→техника"),
        ("software_techniques", "Связей software→техника"),
        ("data_sources", "Источников данных"),
        ("mitigations", "Мер предотвращения"),
    ]

    for table, label in tables:
        count = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
        print(f"  {label}: {count}")

    # Техники с описаниями
    with_desc = conn.execute(
        "SELECT COUNT(*) FROM techniques WHERE description_ru != '' AND description_ru IS NOT NULL"
    ).fetchone()[0]
    total = conn.execute("SELECT COUNT(*) FROM techniques").fetchone()[0]
    print(f"  Техник с RU-описаниями: {with_desc}/{total}")

    print("=" * 50)
    print(f"База сохранена: {DB_PATH}")


def build_catalog(include_stix: bool = False):
    """Основная функция: строит базу из XLSX (и опционально STIX)."""
    if not XLSX_PATH.exists():
        print(f"ОШИБКА: Файл не найден: {XLSX_PATH}")
        print("Скопируй MITRE ATT&CK v18 перевод.xlsm в data/mitre/mitre_v18_translation.xlsm")
        return

    # Удалить старую базу
    if DB_PATH.exists():
        os.remove(DB_PATH)
        print(f"Старая база удалена: {DB_PATH}")

    conn = create_database(DB_PATH)

    try:
        parse_xlsx_tactics_and_techniques(conn, XLSX_PATH)

        if include_stix:
            parse_stix_json(conn, STIX_PATH)

        print_stats(conn)
    finally:
        conn.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="MITRE ATT&CK Catalog Builder")
    parser.add_argument("--stix", action="store_true", help="Также парсить STIX JSON (APT-группы, software)")
    args = parser.parse_args()

    build_catalog(include_stix=args.stix)
